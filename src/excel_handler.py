"""
BP Duplicate Checker - Excel File Handler
==========================================
This module handles reading and writing Excel files (.xlsx)
for the BP Duplicate Checker application.

Features:
- Validate required columns in input files
- Load BP data from Excel
- Export matching results to Excel with formatting
"""

import os
from typing import List, Dict, Tuple, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Required columns for input Excel file
REQUIRED_COLUMNS = ['BP_Number', 'Name1', 'Name2']


class ExcelValidationError(Exception):
    """Custom exception for Excel validation errors."""
    pass


class ExcelHandler:
    """
    Handles all Excel file operations for the BP Duplicate Checker.
    """

    @staticmethod
    def validate_file(file_path: str) -> Tuple[bool, str]:
        """
        Validate that the Excel file exists and has required columns.

        Args:
            file_path: Path to the Excel file

        Returns:
            Tuple of (is_valid, message)
        """
        # Check file exists
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"

        # Check file extension
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return False, "File must be an Excel file (.xlsx or .xls)"

        try:
            # Read only the header row to validate columns
            df = pd.read_excel(file_path, nrows=0)
            columns = [col.strip() for col in df.columns.tolist()]

            # Check for required columns (case-insensitive)
            columns_lower = [col.lower() for col in columns]
            missing = []

            for req_col in REQUIRED_COLUMNS:
                if req_col.lower() not in columns_lower:
                    missing.append(req_col)

            if missing:
                return False, f"Missing required columns: {', '.join(missing)}"

            return True, "File validated successfully"

        except Exception as e:
            return False, f"Error reading file: {str(e)}"

    @staticmethod
    def load_data(file_path: str) -> Tuple[List[Dict[str, str]], str]:
        """
        Load BP data from an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Tuple of (list of records, status message)
        """
        try:
            # Read the Excel file
            df = pd.read_excel(file_path)

            # Standardize column names (handle case variations)
            column_mapping = {}
            for col in df.columns:
                col_lower = col.strip().lower()
                if col_lower == 'bp_number':
                    column_mapping[col] = 'BP_Number'
                elif col_lower == 'name1':
                    column_mapping[col] = 'Name1'
                elif col_lower == 'name2':
                    column_mapping[col] = 'Name2'

            df = df.rename(columns=column_mapping)

            # Convert to list of dictionaries
            # Handle NaN values by converting to empty strings
            df = df.fillna('')
            records = df.to_dict('records')

            return records, f"Loaded {len(records)} records successfully"

        except Exception as e:
            return [], f"Error loading data: {str(e)}"

    @staticmethod
    def export_results(
        results: Dict,
        output_path: str,
        summary_stats: Optional[Dict] = None
    ) -> Tuple[bool, str]:
        """
        Export matching results to a formatted Excel file.

        The output file contains two sheets:
        1. Matching Results - Detailed results with source and match info
        2. Summary - Statistics about the matching process

        Args:
            results: Dictionary of matching results from FuzzyMatcher
            output_path: Path for the output Excel file
            summary_stats: Optional summary statistics to include

        Returns:
            Tuple of (success, message)
        """
        try:
            # Create workbook
            wb = Workbook()

            # ===== Sheet 1: Matching Results =====
            ws_results = wb.active
            ws_results.title = "Matching Results"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            high_score_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            medium_score_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = [
                "Source BP Number",
                "Source Name1",
                "Source Name2",
                "Match Rank",
                "Match BP Number",
                "Match Name1",
                "Match Name2",
                "Similarity Score",
                "Confidence Level"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            # Data rows
            row_num = 2
            for bp_number, matches in results.items():
                if not matches:
                    continue

                for rank, match in enumerate(matches, 1):
                    # Determine confidence level
                    score = match.similarity_score
                    if score >= 80:
                        confidence = "High"
                        score_fill = high_score_fill
                    elif score >= 60:
                        confidence = "Medium"
                        score_fill = medium_score_fill
                    else:
                        confidence = "Low"
                        score_fill = None

                    # Write row data
                    row_data = [
                        match.source_bp.bp_number,
                        match.source_bp.name1,
                        match.source_bp.name2,
                        rank,
                        match.match_bp.bp_number,
                        match.match_bp.name1,
                        match.match_bp.name2,
                        score,
                        confidence
                    ]

                    for col, value in enumerate(row_data, 1):
                        cell = ws_results.cell(row=row_num, column=col, value=value)
                        cell.border = border

                        # Highlight high/medium scores
                        if col == 8 and score_fill:  # Similarity Score column
                            cell.fill = score_fill

                        # Center align certain columns
                        if col in [1, 4, 5, 8, 9]:
                            cell.alignment = Alignment(horizontal='center')

                    row_num += 1

            # Adjust column widths
            column_widths = [15, 25, 25, 10, 15, 25, 25, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws_results.column_dimensions[
                    ws_results.cell(row=1, column=col).column_letter
                ].width = width

            # Freeze header row
            ws_results.freeze_panes = 'A2'

            # ===== Sheet 2: Summary =====
            ws_summary = wb.create_sheet("Summary")

            summary_data = [
                ["BP Duplicate Check - Summary Report", ""],
                ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
            ]

            if summary_stats:
                summary_data.extend([
                    ["Total Records Analyzed", summary_stats.get('total_records', 0)],
                    ["Records with Potential Matches", summary_stats.get('records_with_matches', 0)],
                    ["Total Match Pairs Found", summary_stats.get('total_matches', 0)],
                    ["Average Similarity Score", f"{summary_stats.get('average_score', 0):.2f}%"],
                    ["", ""],
                    ["Confidence Breakdown", ""],
                    ["High Confidence (≥80%)", summary_stats.get('high_confidence', 0)],
                    ["Medium Confidence (60-79%)", summary_stats.get('medium_confidence', 0)],
                    ["Low Confidence (<60%)", summary_stats.get('low_confidence', 0)],
                ])

            for row_num, (label, value) in enumerate(summary_data, 1):
                cell_label = ws_summary.cell(row=row_num, column=1, value=label)
                cell_value = ws_summary.cell(row=row_num, column=2, value=value)

                if row_num == 1:
                    cell_label.font = Font(bold=True, size=14)
                elif label and not value:
                    cell_label.font = Font(bold=True)

            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 25

            # Save workbook
            wb.save(output_path)
            return True, f"Results exported to: {output_path}"

        except PermissionError:
            return False, "Cannot save file - it may be open in another application"
        except Exception as e:
            return False, f"Error exporting results: {str(e)}"


def create_example_input_file(output_path: str) -> Tuple[bool, str]:
    """
    Create an example input Excel file for testing.

    Args:
        output_path: Path for the example file

    Returns:
        Tuple of (success, message)
    """
    try:
        # Example data with various name variations
        example_data = [
            {"BP_Number": "BP001", "Name1": "ABC Company Ltd.", "Name2": ""},
            {"BP_Number": "BP002", "Name1": "ABC Company Limited", "Name2": ""},
            {"BP_Number": "BP003", "Name1": "Mrs. Jane Smith", "Name2": ""},
            {"BP_Number": "BP004", "Name1": "Jane Smith", "Name2": ""},
            {"BP_Number": "BP005", "Name1": "XYZ Corporation", "Name2": "Technology Division"},
            {"BP_Number": "BP006", "Name1": "XYZ Corp.", "Name2": "Tech Division"},
            {"BP_Number": "BP007", "Name1": "Global Trading Co.", "Name2": ""},
            {"BP_Number": "BP008", "Name1": "Global Trading Company", "Name2": "International"},
            {"BP_Number": "BP009", "Name1": "Mr. John Doe", "Name2": ""},
            {"BP_Number": "BP010", "Name1": "John Doe", "Name2": "Senior Partner"},
            {"BP_Number": "BP011", "Name1": "Smith & Associates LLC", "Name2": ""},
            {"BP_Number": "BP012", "Name1": "Smith and Associates", "Name2": "LLC"},
            {"BP_Number": "BP013", "Name1": "First National Bank", "Name2": ""},
            {"BP_Number": "BP014", "Name1": "1st National Bank", "Name2": ""},
            {"BP_Number": "BP015", "Name1": "Acme Industries", "Name2": ""},
            {"BP_Number": "BP016", "Name1": "ACME INDUSTRIES LTD", "Name2": ""},
            {"BP_Number": "BP017", "Name1": "Tech Solutions Inc.", "Name2": ""},
            {"BP_Number": "BP018", "Name1": "Unique Company", "Name2": "No Matches"},
            {"BP_Number": "BP019", "Name1": "Pacific Trading", "Name2": "Enterprises"},
            {"BP_Number": "BP020", "Name1": "Pacific Trading Enterprises", "Name2": ""},
        ]

        df = pd.DataFrame(example_data)
        df.to_excel(output_path, index=False)

        return True, f"Example file created: {output_path}"

    except Exception as e:
        return False, f"Error creating example file: {str(e)}"
